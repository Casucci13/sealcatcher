import os
import signal

from flask import (
    Flask,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    url_for,
)
from openpyxl import Workbook, load_workbook
from replit import db

app = Flask(__name__)



def copy_excel_file(src_path='dataPP.xlsx', dest_path='dataCopy.xlsx'):
    # Load the existing workbook
    workbook = load_workbook(src_path)
    
    # Save the copy as a new file
    workbook.save(dest_path)
    
    print(f"Copied {src_path} to {dest_path}")


def handle_signal(signum, frame):
    print(f"Received signal {signum}, performing graceful shutdown...")
    cleanup_function()
    sys.exit(0)
signal.signal(signal.SIGTERM, handle_signal)
signal.signal(signal.SIGINT, handle_signal)

def cleanup_function(file_path='dataPP.xlsx'):
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # Perform any necessary final updates to the workbook here
        # Example: You might want to add a final timestamp, summary, etc.
        # Save the workbook
        
        workbook.save(file_path)
        copy_excel_file()
        
        print(f"Workbook {file_path} successfully saved.")
    except Exception as e:
        print(f"An error occurred while saving {file_path}: {e}")

@app.route('/')
def home():
    # Load the Excel files
    datapp_wb = load_workbook('dataPP.xlsx')
    datatags_wb = load_workbook('datatags.xlsx')
    datapp_sheet = datapp_wb.active
    datatags_sheet = datatags_wb.active
    # Extract and count unique pad_tags
    unique_pad_tags = set()
    for row in datapp_sheet.iter_rows(values_only=True):
        if row[0] is not None:  # Assuming pad_tag is in the 4th column
            unique_pad_tags.add(row[0])
    unique_pad_tag_count = len(unique_pad_tags)
    # Count rows with text in datatags.xlsx
    text_rows = 0
    for row in datatags_sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            text_rows += 1
    # Calculate percentage
    percentage = (unique_pad_tag_count / text_rows) * 100
    # Pass the percentage to the template
    return render_template('home.html', percentage=percentage, unique_pad_tag_count=unique_pad_tag_count, text_rows=text_rows)

    # Load the Excel files
    

@app.route('/search_and_select', methods=['GET', 'POST'])
def search_and_select():
    return render_template('search_and_select.html')

@app.route('/search', methods=['GET'])
def search():
    search_value = request.args.get('search_value')
    results = []
    tag_description = None
    file_path3 = 'data3.xlsx'
    file_path_tags = 'datatags.xlsx'

    if search_value:
        if os.path.exists(file_path3):
            workbook3 = load_workbook(file_path3)
            sheet3 = workbook3.active
            for row in sheet3.iter_rows(values_only=True):
                if str(row[1]).lower() == str(search_value).lower():  # Case insensitive comparison
                    results.append(row[0])

        if os.path.exists(file_path_tags):
            workbook_tags = load_workbook(file_path_tags)
            sheet_tags = workbook_tags.active
            for row in sheet_tags.iter_rows(values_only=True):
                if str(row[0]).lower() == str(search_value).lower():
                    tag_description = row[1]
                    break
        
    return jsonify({'results': results, 'tag_description': tag_description})

@app.route('/submit_data', methods=['POST'])
def submit_data():
    search_value = request.form.get('search_value')
    entered_values = request.form.getlist('ED')
    case_number = request.form.get('case_number')
    ipad_number = request.form.get('ipad_number')
    cradle_point_serial = request.form.get('cradle_point_serial')
    PP = 0
    
    file_path_pp = 'dataPP.xlsx'
    file_path_2 = 'data2.xlsx'
    
    
    if not os.path.exists(file_path_pp):
        workbook_pp = Workbook()
        sheet_pp = workbook_pp.active
        
    else:
        workbook_pp = load_workbook(file_path_pp)
        sheet_pp = workbook_pp.active

    for value in entered_values:
        sheet_pp.append([search_value, value, PP, case_number, ipad_number, cradle_point_serial])
        # Read data2.xlsx
        workbook_2 = load_workbook(file_path_2)
        sheet_2 = workbook_2.active
        # Iterate through dataPP.xlsx
        for row_index, row in enumerate(sheet_pp.iter_rows(values_only=True),            start=1):
            if row_index > 0: # Skip header row
                value_to_match = row[1] # Value from 2nd column of dataPP.xlsx
                # Find match in data2.xlsx
                for row_2 in sheet_2.iter_rows(values_only=True):
                    if row_2[0] == value_to_match: # Match found
                        related_value = row_2[1] # Value from 2nd column of data2.xlsx
                        sheet_pp.cell(row=row_index, column=3).value = related_value # Append to 3rd column of dataPP.xlsx
    workbook_pp.save(file_path_pp)
    return redirect(url_for('search_and_select'))

@app.route('/download_data')
def download_data():
        copy_excel_file()
        
        return send_from_directory(directory=os.getcwd(), path='dataPP.xlsx', as_attachment=True)

@app.route('/enter_values', methods=['GET', 'POST'])
def enter_values():
    entries = []
    value1 = None
    value2 = None
    value3 = None
    related_value = None
    file_path = 'data.xlsx'
    file_path2 = 'data2.xlsx'

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        entries = list(sheet.iter_rows(values_only=True))[-20:]  # Load last 20 rows to display recent 10 entries (2 rows per entry)
        entries.reverse()

    if request.method == 'POST':
        value1 = request.form.get('value1')
        value2 = request.form.get('value2')
        value3 = request.form.get('value3')

        if value2 and value3:
            if not workbook:
                workbook = Workbook()
                sheet = workbook.active

            # Append the new values to the sheet
            sheet.append([value1, value2])
            sheet.append([value1, value3])
            workbook.save(file_path)

            # Reload the entries after adding the new entry
            entries = list(sheet.iter_rows(values_only=True))[-20:]
            entries.reverse()

            return redirect(url_for('enter_values'))

        # If only value1 is provided, find the corresponding value2 and value3
        if value1:
            for i in range(0, len(entries), 2):
                if str(entries[i][0]).lower() == str(value1).lower():
                    value2 = entries[i][1]
                    value3 = entries[i + 1][1]
                    break

            # Fetch related value from data2.xlsx
            if os.path.exists(file_path2):
                workbook2 = load_workbook(file_path2)
                sheet2 = workbook2.active
                for row in sheet2.iter_rows(values_only=True):
                    if str(row[0]).lower() == str(value1).lower():
                        related_value = row[1]
                        break

    # Organize entries for display
    display_entries = []
    for i in range(0, len(entries), 2):
        if i + 1 < len(entries):
            display_entries.append(entries[i][0])

    return render_template('enter_values.html', entries=display_entries, value1=value1, value2=value2, value3=value3, related_value=related_value)

@app.route('/get_related_value', methods=['GET'])
def get_related_value():
    value1 = request.args.get('value1')
    related_value = None
    file_path2 = 'data2.xlsx'

    if value1 and os.path.exists(file_path2):
        workbook2 = load_workbook(file_path2)
        sheet2 = workbook2.active
        for row in sheet2.iter_rows(values_only=True):
            if str(row[0]).lower() == str(value1).lower():
                related_value = row[1]
                break

    return jsonify({'related_value': related_value})

@app.route('/enter_11_values', methods=['GET', 'POST'])
def enter_11_values():
    entries = []
    file_path = 'data_11.xlsx'

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        entries = list(sheet.iter_rows(values_only=True))[-10:]
        entries.reverse()

    if request.method == 'POST':
        values = [request.form[f'value{i}'] for i in range(1, 12)]

        if not workbook:
            workbook = Workbook()
            sheet = workbook.active

        sheet.append(values)
        workbook.save(file_path)

        entries = list(sheet.iter_rows(values_only=True))[-10:]
        entries.reverse()

        return redirect(url_for('enter_11_values'))

    return render_template('enter_11_values.html', entries=entries)

@app.route('/another_page')
def another_page():
    data_entries = []
    data_11_entries = []

    if os.path.exists('dataPP.xlsx'):
        workbook = load_workbook('dataPP.xlsx')
        sheet = workbook.active
        data_entries = list(sheet.iter_rows(values_only=True))[-20:]  # Load last 20 rows
    data_entries.reverse()  # Reverse to show recent entries first
    datapp_wb = load_workbook('dataPP.xlsx')
    datatags_wb = load_workbook('datatags.xlsx')
    datapp_sheet = datapp_wb.active
    datatags_sheet = datatags_wb.active
    # Extract and count unique pad_tags
    unique_pad_tags = set()
    for row in datapp_sheet.iter_rows(values_only=True):
        if row[0] is not None:  # Assuming pad_tag is in the 4th column
            unique_pad_tags.add(row[0])
    unique_pad_tag_count = len(unique_pad_tags)
    # Count rows with text in datatags.xlsx
    text_rows = 0
    for row in datatags_sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            text_rows += 1
    # Calculate percentage
    percentage = ((unique_pad_tag_count) / text_rows) * 100
    # Pass the percentage to the template
    return render_template('another_page.html', percentage=percentage, unique_pad_tag_count=unique_pad_tag_count, text_rows=text_rows, data_entries=data_entries, data_11_entries=data_11_entries)


   # return render_template('another_page.html', data_entries=data_entries, data_11_entries=data_11_entries)
    
@app.route('/reset_database')
def reset_database():
    file_path_pp = 'dataPP.xlsx'
    if os.path.exists(file_path_pp):
        workbook_pp = Workbook()
        sheet_pp = workbook_pp.active
        
        workbook_pp.save(file_path_pp)
    return jsonify({'success': True})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)





